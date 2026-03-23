"""
AstroNifty — Daily Data Exporter (Multi-User).

Per-user exports: each user gets their own directory under
    exports/{user_id}/{YYYYMMDD}/

Contains CSVs, consolidated Excel workbook, and PDF report
with ONLY that user's trades (not all users' trades).
"""

import os
from datetime import datetime, date
from pathlib import Path
from typing import List, Optional

import pandas as pd
from loguru import logger

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    logger.warning("openpyxl not installed. Run: pip install openpyxl")
    raise


EXPORTS_DIR = Path(__file__).parent.parent.parent / "exports"


class DailyExporter:
    """
    Per-user daily data exporter.

    Creates CSV files per data type and a consolidated Excel workbook.
    Each user's exports are isolated in their own directory:
        exports/{user_id}/{YYYYMMDD}/
    """

    def __init__(self, exports_dir: Optional[str] = None):
        """
        Args:
            exports_dir: Root directory for exports. Defaults to project/exports/.
        """
        self.exports_dir = Path(exports_dir) if exports_dir else EXPORTS_DIR
        self.exports_dir.mkdir(parents=True, exist_ok=True)
        logger.info("DailyExporter initialized. Exports dir: {}", self.exports_dir)

    # ──────────────────────────────────────────────────────────────────────
    # Main daily export (runs at 4 PM, per user)
    # ──────────────────────────────────────────────────────────────────────

    def export_daily(self, db_manager, user_id: str) -> dict:
        """
        Run the full daily export pipeline for a SPECIFIC user.

        Creates per-user directory: exports/{user_id}/{YYYYMMDD}/
        Exports only that user's trades and signals.
        Shared data (OI, astro, FII/DII) is included as context.

        Args:
            db_manager: Database manager instance with query methods.
            user_id:    The user to export data for.

        Returns:
            dict with export paths and record counts.
        """
        today_str = date.today().strftime("%Y%m%d")
        # Per-user export directory
        day_dir = self.exports_dir / user_id / today_str
        day_dir.mkdir(parents=True, exist_ok=True)
        logger.info(
            "[USER:{}] Starting daily export for {} -> {}",
            user_id,
            today_str,
            day_dir,
        )

        result = {
            "user_id": user_id,
            "date": today_str,
            "directory": str(day_dir),
            "files": {},
            "record_counts": {},
            "errors": [],
        }

        # ── Fetch shared data (same for all users) ──
        shared_datasets = {}
        shared_fetch_map = {
            "oi_chain": "get_todays_oi_snapshots",
            "astro": "get_todays_astro",
            "fii_dii": "get_todays_fii_dii",
            "weekly_forecast": "get_weekly_forecast",
        }

        for key, method_name in shared_fetch_map.items():
            try:
                method = getattr(db_manager, method_name, None)
                if method is None:
                    shared_datasets[key] = []
                    result["errors"].append(f"Missing method: {method_name}")
                    continue
                raw = method()
                shared_datasets[key] = raw if isinstance(raw, list) else ([raw] if raw else [])
            except Exception as e:
                logger.error("[USER:{}] Error fetching {}: {}", user_id, key, str(e))
                shared_datasets[key] = []
                result["errors"].append(f"Fetch error for {key}: {str(e)}")

        # ── Fetch per-user data ──
        user_datasets = {}
        user_fetch_map = {
            "signals": ("get_todays_signals", {"user_id": user_id}),
            "trades": ("get_todays_trades", {"user_id": user_id}),
            "pnl": ("get_todays_pnl", {"user_id": user_id}),
        }

        for key, (method_name, kwargs) in user_fetch_map.items():
            try:
                method = getattr(db_manager, method_name, None)
                if method is None:
                    user_datasets[key] = []
                    result["errors"].append(f"Missing method: {method_name}")
                    continue
                try:
                    raw = method(**kwargs)
                except TypeError:
                    # Fallback if method doesn't accept user_id kwarg
                    raw = method()
                user_datasets[key] = raw if isinstance(raw, list) else ([raw] if raw else [])
            except Exception as e:
                logger.error("[USER:{}] Error fetching {}: {}", user_id, key, str(e))
                user_datasets[key] = []
                result["errors"].append(f"Fetch error for {key}: {str(e)}")

        # Merge datasets for export
        datasets = {**shared_datasets, **user_datasets}

        # ── Export CSVs (shared context) ──
        shared_csv_map = {
            "OI_CHAIN.csv": "oi_chain",
            "ASTRO.csv": "astro",
            "FII_DII.csv": "fii_dii",
            "WEEKLY_FORECAST.csv": "weekly_forecast",
        }

        for filename, data_key in shared_csv_map.items():
            try:
                filepath = day_dir / filename
                records = datasets.get(data_key, [])
                if records:
                    df = pd.DataFrame(records)
                    df.to_csv(filepath, index=False, encoding="utf-8")
                    result["files"][filename] = str(filepath)
                    result["record_counts"][data_key] = len(records)
                    logger.info("[USER:{}] Exported {} ({} rows)", user_id, filename, len(records))
                else:
                    filepath.write_text("No data for this date\n", encoding="utf-8")
                    result["files"][filename] = str(filepath)
                    result["record_counts"][data_key] = 0
            except Exception as e:
                logger.error("[USER:{}] CSV export failed for {}: {}", user_id, filename, str(e))
                result["errors"].append(f"CSV error for {filename}: {str(e)}")

        # ── Export per-user signals CSV ──
        try:
            signals = datasets.get("signals", [])
            signals_path = day_dir / "SIGNALS.csv"
            if signals:
                df = pd.DataFrame(signals)
                df.to_csv(signals_path, index=False, encoding="utf-8")
            else:
                signals_path.write_text("No signals for this date\n", encoding="utf-8")
            result["files"]["SIGNALS.csv"] = str(signals_path)
            result["record_counts"]["signals"] = len(signals)
        except Exception as e:
            logger.error("[USER:{}] Signals CSV export failed: {}", user_id, str(e))
            result["errors"].append(f"Signals CSV error: {str(e)}")

        # ── Export per-user trade log CSV ──
        try:
            trades = datasets.get("trades", [])
            trade_csv_path = day_dir / "TRADE_LOG.csv"
            if trades:
                trade_csv = self.export_trade_log(trades)
                trade_csv_path.write_text(trade_csv, encoding="utf-8")
            else:
                trade_csv_path.write_text("No trades for this date\n", encoding="utf-8")
            result["files"]["TRADE_LOG.csv"] = str(trade_csv_path)
            result["record_counts"]["trades"] = len(trades)
        except Exception as e:
            logger.error("[USER:{}] Trade log CSV export failed: {}", user_id, str(e))
            result["errors"].append(f"Trade log CSV error: {str(e)}")

        # ── Consolidated Excel Workbook (per user) ──
        try:
            xlsx_name = f"ASTRONIFTY_DAILY_{user_id}_{today_str}.xlsx"
            xlsx_path = day_dir / xlsx_name
            self._build_excel_workbook(xlsx_path, datasets, today_str, user_id)
            result["files"][xlsx_name] = str(xlsx_path)
            logger.info("[USER:{}] Excel workbook exported: {}", user_id, xlsx_path)
        except Exception as e:
            logger.error("[USER:{}] Excel export failed: {}", user_id, str(e))
            result["errors"].append(f"Excel error: {str(e)}")

        # ── PDF Report (per user — only their trades) ──
        try:
            from database.pdf_exporter import DailyPDFExporter

            pdf_exporter = DailyPDFExporter(exports_dir=str(day_dir))

            engine_data = {
                "oi_data": datasets.get("oi_chain", {}),
                "astro_data": datasets.get("astro", {}),
                "signals": datasets.get("signals", []),
                "fii_dii_data": datasets.get("fii_dii", {}),
                "weekly_forecast": datasets.get("weekly_forecast", {}),
                "trades": datasets.get("trades", []),
                "pnl_summary": datasets.get("pnl", {}),
                "user_id": user_id,
            }

            for key, method_name in [
                ("scores", "get_todays_scores"),
                ("greeks_data", "get_todays_greeks"),
                ("global_cues", "get_global_cues"),
                ("sector_data", "get_todays_sectors"),
                ("price_action", "get_price_action"),
                ("risk_summary", "get_risk_summary"),
                ("market_summary", "get_market_summary"),
            ]:
                try:
                    method = getattr(db_manager, method_name, None)
                    if method:
                        engine_data[key] = method() or {}
                except Exception:
                    engine_data[key] = {}

            pdf_path = pdf_exporter.generate_daily_report(
                report_date=date.today(),
                engine_data=engine_data,
            )
            pdf_name = f"ASTRONIFTY_DAILY_{user_id}_{today_str}.pdf"
            result["files"][pdf_name] = pdf_path
            logger.info("[USER:{}] PDF report exported: {}", user_id, pdf_path)
        except Exception as e:
            logger.error("[USER:{}] PDF export failed: {}", user_id, str(e))
            result["errors"].append(f"PDF error: {str(e)}")

        # ── Save export record to database ──
        try:
            export_record = {
                "user_id": user_id,
                "date": today_str,
                "timestamp": datetime.now().isoformat(),
                "directory": str(day_dir),
                "files": list(result["files"].keys()),
                "record_counts": result["record_counts"],
                "errors": result["errors"],
                "status": "success" if not result["errors"] else "partial",
            }
            if hasattr(db_manager, "save_export_record"):
                db_manager.save_export_record(export_record)
                logger.info("[USER:{}] Export record saved to database", user_id)
        except Exception as e:
            logger.error("[USER:{}] Failed to save export record: {}", user_id, str(e))

        logger.info(
            "[USER:{}] Daily export complete: {} files, {} errors",
            user_id,
            len(result["files"]),
            len(result["errors"]),
        )
        return result

    # ──────────────────────────────────────────────────────────────────────
    # Trade log export
    # ──────────────────────────────────────────────────────────────────────

    def export_trade_log(self, trades: list) -> str:
        """Export a list of trade dicts to CSV string."""
        if not trades:
            return "No trades\n"

        df = pd.DataFrame(trades)

        preferred_cols = [
            "time", "date", "index", "instrument", "direction", "qty",
            "entry_price", "exit_price", "sl", "target", "pnl",
            "pnl_pct", "reason", "signal_score", "status", "user_id",
        ]
        ordered_cols = [c for c in preferred_cols if c in df.columns]
        remaining = [c for c in df.columns if c not in ordered_cols]
        df = df[ordered_cols + remaining]

        csv_str = df.to_csv(index=False)
        logger.debug("Trade log CSV generated: {} trades, {} bytes", len(trades), len(csv_str))
        return csv_str

    # ──────────────────────────────────────────────────────────────────────
    # OI snapshots export
    # ──────────────────────────────────────────────────────────────────────

    def export_oi_snapshots(self, snapshots: list) -> str:
        """Export OI chain snapshots to CSV string."""
        if not snapshots:
            return "No OI snapshots\n"

        df = pd.DataFrame(snapshots)

        preferred_cols = [
            "timestamp", "index", "expiry", "strike",
            "ce_oi", "pe_oi", "ce_change", "pe_change",
            "ce_volume", "pe_volume", "ce_iv", "pe_iv",
            "ce_ltp", "pe_ltp", "spot", "max_pain", "pcr",
        ]
        ordered_cols = [c for c in preferred_cols if c in df.columns]
        remaining = [c for c in df.columns if c not in ordered_cols]
        df = df[ordered_cols + remaining]

        csv_str = df.to_csv(index=False)
        logger.debug("OI snapshots CSV generated: {} rows, {} bytes", len(snapshots), len(csv_str))
        return csv_str

    # ──────────────────────────────────────────────────────────────────────
    # Excel workbook builder (per user)
    # ──────────────────────────────────────────────────────────────────────

    def _build_excel_workbook(
        self,
        filepath: Path,
        datasets: dict,
        date_str: str,
        user_id: str,
    ):
        """Build a consolidated Excel workbook with styled sheets."""
        wb = openpyxl.Workbook()

        # Style definitions
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        green_font = Font(color="00C853")
        red_font = Font(color="FF1744")
        border = Border(
            left=Side(style="thin", color="333333"),
            right=Side(style="thin", color="333333"),
            top=Side(style="thin", color="333333"),
            bottom=Side(style="thin", color="333333"),
        )
        center_align = Alignment(horizontal="center", vertical="center")

        def style_header(ws):
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border

        def auto_width(ws):
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        val_len = len(str(cell.value or ""))
                        if val_len > max_len:
                            max_len = val_len
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

        def add_sheet(name: str, data_key: str):
            records = datasets.get(data_key, [])
            if not records:
                ws = wb.create_sheet(title=name)
                ws["A1"] = f"No {name} data for {date_str}"
                ws["A1"].font = Font(italic=True, color="888888")
                return
            df = pd.DataFrame(records)
            ws = wb.create_sheet(title=name)
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = border
                    cell.alignment = center_align
            style_header(ws)
            auto_width(ws)

            pnl_cols = [
                i for i, col_name in enumerate(df.columns, 1)
                if "pnl" in col_name.lower()
            ]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for col_idx in pnl_cols:
                    if col_idx <= len(row):
                        cell = row[col_idx - 1]
                        try:
                            val = float(cell.value or 0)
                            cell.font = green_font if val >= 0 else red_font
                        except (ValueError, TypeError):
                            pass

        # ── Summary sheet (default) ──
        summary_ws = wb.active
        summary_ws.title = "Summary"
        summary_data = [
            ["ASTRONIFTY DAILY REPORT", ""],
            ["User", user_id],
            ["Date", date_str],
            ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            [""],
            ["Dataset", "Records"],
        ]
        for key, records in datasets.items():
            count = len(records) if isinstance(records, list) else (1 if records else 0)
            summary_data.append([key.upper().replace("_", " "), count])

        for r_idx, row in enumerate(summary_data, 1):
            for c_idx, value in enumerate(row, 1):
                cell = summary_ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border

        summary_ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="00E5FF")
        summary_ws.column_dimensions["A"].width = 30
        summary_ws.column_dimensions["B"].width = 25

        # ── Data sheets ──
        add_sheet("OI Chain", "oi_chain")
        add_sheet("Astro", "astro")
        add_sheet("Signals", "signals")
        add_sheet("FII DII", "fii_dii")
        add_sheet("Weekly Forecast", "weekly_forecast")
        add_sheet("Trades", "trades")

        # ── P&L sheet ──
        pnl_data = datasets.get("pnl", [])
        if pnl_data:
            pnl_ws = wb.create_sheet(title="P&L")
            if isinstance(pnl_data, list) and len(pnl_data) > 0:
                item = pnl_data[0] if isinstance(pnl_data[0], dict) else {}
            elif isinstance(pnl_data, dict):
                item = pnl_data
            else:
                item = {}

            row_idx = 1
            pnl_ws.cell(row=row_idx, column=1, value=f"P&L SUMMARY — {user_id}").font = Font(
                size=13, bold=True, color="00E5FF"
            )
            row_idx += 1
            for k, v in item.items():
                pnl_ws.cell(row=row_idx, column=1, value=str(k).upper()).font = Font(bold=True)
                cell = pnl_ws.cell(row=row_idx, column=2, value=v)
                try:
                    val = float(v)
                    if "pnl" in k.lower() or "profit" in k.lower() or "loss" in k.lower():
                        cell.font = green_font if val >= 0 else red_font
                except (ValueError, TypeError):
                    pass
                row_idx += 1
            pnl_ws.column_dimensions["A"].width = 25
            pnl_ws.column_dimensions["B"].width = 20

        wb.save(filepath)
        logger.info(
            "[USER:{}] Excel workbook saved: {} ({} sheets)",
            user_id,
            filepath,
            len(wb.sheetnames),
        )
